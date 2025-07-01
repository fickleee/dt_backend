import os
import openpyxl
import datetime
import time
import sqlite3

# 数据库连接
conn = sqlite3.connect('./database/datang.db')
cursor = conn.cursor()
station_name = "datu"
# 遍历rawdata下的所有文件夹
data_cache_dir = 'dataCache'
column_mapping = {
    "序号": "index",
    "Time": "timestamp",
    "YC02127": "irr1",
    "YC02130": "irr2"
}
for root, dirs, files in os.walk(data_cache_dir):
    print(files)
    for file in files:
        print(file)
        if file.endswith('.xlsx'):
            # 解析文件名以获取box_id和inverter_id
            file_name = file.split('.xlsx')[0]

            # 读取Excel文件
            workbook = openpyxl.load_workbook(os.path.join(root, file))
            sheet = workbook.active
            
            # 处理第一列前两个单元格合并，并删除第二行
            first_cell = sheet.cell(row=1, column=1).value
            first_cell_split = first_cell.split(' ') if first_cell else []
            sheet.delete_rows(2)  # 删除第二行
            
            headers = [cell.value for cell in sheet[1]]
            # 写入处理后的数据到数据库
            for row in sheet.iter_rows(min_row=2, values_only=True):
                obj = {}
                for colid in range(len(row)):
                    col_name = headers[colid]
                    key_name = column_mapping[col_name]
                    obj[key_name] = row[colid]

                time_str = obj["timestamp"]
                timestamp = int(time.mktime(datetime.datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S").timetuple()))
                obj["timestamp"] = timestamp

                start_date = int(time.mktime(datetime.datetime.strptime('2024-01-01 00:00:00', '%Y-%m-%d %H:%M:%S').timetuple()))
                end_date = int(time.mktime(datetime.datetime.strptime('2024-10-31 23:59:59', '%Y-%m-%d %H:%M:%S').timetuple()))
                if timestamp < start_date or timestamp > end_date:
                    continue
                
                inverterQuery = '''
                INSERT INTO {}StationInfo (
                    timestamp, irradiance, temperature, power
                ) VALUES (?, ?, ?, ?)
                '''.format(station_name.lower().replace(' ', '_'))
                print(inverterQuery)
                try:
                    cursor.execute(inverterQuery,(
                        obj["timestamp"], (obj["irr1"] + obj["irr2"]) / 2, 25, 0 
                    ))
                except:
                    pass
# 提交事务
conn.commit()

# 关闭Cursor和Connection
cursor.close()
conn.close()