import os
import openpyxl
import datetime
import time
import sqlite3

# Get the current directory path
current_dir = os.path.dirname(os.path.abspath(__file__))

# 数据库连接
conn = sqlite3.connect(os.path.join(current_dir, 'database', 'datang.db'))
cursor = conn.cursor()
station_name = "datu"
# 遍历rawdata下的所有文件夹
data_cache_dir = os.path.join(current_dir, 'rawdata')


reverse_column_mapping = {
    "index": "序号",
    "timestamp": "Time",
    "temperature": "PvTmp",
    "sig_overvoltage": "YX00001",
    "sig_undervoltage": "YX00002",
    "sig_overfrequency": "YX00003",
    "sig_underfrequency": "YX00004",
    "sig_gridless": "YX00006",
    "sig_imbalance": "YX00007",
    "sig_overcurrent": "YX00009",
    "sig_midpoint_grounding": "YX00021",
    "sig_insulation_failure": "YX00026",
    "sig_excessive_DC": "YX00030",
    "sig_arc_self_protection": "YX00034",
    "sig_arc_failure": "YX00035",
    "intensity": "PVINV_DCI",
    "power": "PVINV_DCPWR",
    "voltage": "PVINV_DCV",
    "generated_energy": "TotWhD2"
}

# 对于StringInfo中的intensity和voltage列，我们需要动态生成键值对
for i in range(1, 31):  # 假设有30个string_id，从1到30
    reverse_column_mapping[f"string_info_intensity_{i}"] = f"PVINV_DCI{i:02d}"
    reverse_column_mapping[f"string_info_voltage_{i}"] = f"PVINV_DCV{i:02d}"

column_mapping = {
    "序号": "index",
    "Time": "timestamp",
    "PvTmp": "temperature",
    "YX00001": "sig_overvoltage",
    "YX00002": "sig_undervoltage",
    "YX00003": "sig_overfrequency",
    "YX00004": "sig_underfrequency",
    "YX00006": "sig_gridless",
    "YX00007": "sig_imbalance",
    "YX00009": "sig_overcurrent",
    "YX00021": "sig_midpoint_grounding",
    "YX00026": "sig_insulation_failure",
    "YX00030": "sig_excessive_DC",
    "YX00034": "sig_arc_self_protection",
    "YX00035": "sig_arc_failure",
    "PVINV_DCI": "intensity",
    "PVINV_DCPWR": "power",
    "PVINV_DCV": "voltage",
    "TotWhD2": "generated_energy"
}

# 对于"PVINV_DCI{string_id}"和"PVINV_DCV{string_id}"的列头，我们需要动态生成键值对
for i in range(1, 31):  # 假设有30个string_id，从1到30
    column_mapping[f"PVINV_DCI{i:02d}"] = f"string_info_intensity_{i}"
    column_mapping[f"PVINV_DCV{i:02d}"] = f"string_info_voltage_{i}"

for root, dirs, files in os.walk(data_cache_dir):
    
    rec = ""
    print(dirs)
    for dir in dirs:
        cnt = 0
        print(dir)
        files = os.listdir(data_cache_dir+"/"+dir)
        for file in files:
            if file.startswith('DTZJJK-CDTGF-Q1-BT') and file.endswith('.xlsx'):
                cnt += 1
                # 解析文件名以获取box_id和inverter_id
                file_name = file.split('.xlsx')[0]
                box_id = file_name.split('Q')[-1].split('-')[1][2:].zfill(3)
                inverter_id = file_name.split('Q')[-1].split('-')[-1][1:].zfill(3)

                # 读取Excel文件
                workbook = openpyxl.load_workbook(os.path.join(root, dir+"/"+file))
                sheet = workbook.active
                
                # 处理第一列前两个单元格合并，并删除第二行
                first_cell = sheet.cell(row=1, column=1).value
                first_cell_split = first_cell.split(' ') if first_cell else []
                sheet.delete_rows(2)  # 删除第二行
                
                headers = [cell.value for cell in sheet[1]]
                # 写入处理后的数据到数据库
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    obj = {}
                    for colid in range(len(row)):
                        col_name = headers[colid]
                        key_name = column_mapping[col_name]
                        obj[key_name] = row[colid]

                    time_str = obj["timestamp"]
                    timestamp = int(time.mktime(datetime.datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S").timetuple()))
                    obj["timestamp"] = timestamp

                    start_date = int(time.mktime(datetime.datetime.strptime('2024-01-01 00:00:00', '%Y-%m-%d %H:%M:%S').timetuple()))
                    end_date = int(time.mktime(datetime.datetime.strptime('2024-10-31 23:59:59', '%Y-%m-%d %H:%M:%S').timetuple()))
                    if timestamp < start_date or timestamp > end_date:
                        continue
                    device_id_inverter = f"{box_id}-{inverter_id}"
                    inverterQuery = '''
                    INSERT INTO {}InverterInfo (
                        timestamp, device_id, inverter_id, box_id, intensity, voltage, power,
                        temperature, sig_overvoltage, sig_undervoltage, sig_overfrequency,
                        sig_underfrequency, sig_gridless, sig_imbalance, sig_overcurrent,
                        sig_midpoint_grounding, sig_insulation_failure, sig_excessive_DC,
                        sig_arc_self_protection, sig_arc_failure, generated_energy
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    '''.format(station_name.lower().replace(' ', '_'))
                    try:
                        cursor.execute(inverterQuery,(
                            obj["timestamp"], device_id_inverter, inverter_id, box_id, obj["intensity"], obj["voltage"], obj["power"],
                            obj["temperature"], obj["sig_overvoltage"], obj["sig_undervoltage"], obj["sig_overfrequency"],
                            obj["sig_underfrequency"], obj["sig_gridless"], obj["sig_imbalance"], obj["sig_overcurrent"],
                            obj["sig_midpoint_grounding"], obj["sig_insulation_failure"], obj["sig_excessive_DC"],
                            obj["sig_arc_self_protection"], obj["sig_arc_failure"], obj["generated_energy"]  # 假设generated_energy为0，因为没有提供
                        ))
                    except:
                        pass
                    # 插入StringInfo数据
                    for i in range(1,31):
                        string_id = str(i).zfill(3)
                        device_id_string = f"{device_id_inverter}-{string_id}"
                        if f"string_info_voltage_{i}" not in obj:
                            obj[f"string_info_voltage_{i}"] = None
                        query = '''
                        INSERT INTO {}StringInfo (
                            timestamp, device_id, string_id, inverter_id, box_id, intensity,
                            voltage, fixed_intensity, fixed_voltage
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        '''.format(station_name.lower().replace(' ', '_'))
                        try:
                            cursor.execute(query, (
                                obj["timestamp"], device_id_string, string_id, inverter_id, box_id, obj[f"string_info_intensity_{i}"], obj[f"string_info_voltage_{i}"], None, None
                            ))
                        except:
                            pass
            rec = file
            conn.commit()
            print("submitted file", file)
        if cnt != len(files):
            print("miss some files in the folder", dir, file)
# 提交事务
conn.commit()

# 关闭Cursor和Connection
cursor.close()
conn.close()