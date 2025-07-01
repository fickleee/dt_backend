import json
import os
from datetime import datetime
import time
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

# 添加日志函数
def silent_log(*args, **kwargs):
    # 如果需要调试，可以取消下面一行的注释
    print(*args, **kwargs)

# 场站名称映射（英文到中文）
STATION_NAME_MAPPING = {
    'datu': '长大涂',
    'wushashan': '乌沙山',
    'fuyang': '富阳',
    'daxue': '大峃',
    'eryuan': '二源',
    'tangjing': '唐景',
    'tangyun': '唐云',
    'mayu': '马屿'
}

CENTER_NAME_MAPPING = {
    'hangzhou': '杭州运维中心',
    'ningbo': '宁波运维中心',
    'wenzhou': '温州运维中心',
    'lishui': '丽水运维中心',
}

CENTER_TO_STATION = {
    'hangzhou': ['fuyang'],
    'ningbo': ['datu', 'wushashan'],
    'wenzhou': ['daxue', 'eryuan', 'mayu'],
    'lishui': ['tangjing', 'tangyun'],
}

# 告警名称映射
INVERTER_ALARM_MAPPING = {
    "sig_arc_failure": "直流电弧故障",
    "sig_arc_self_protection": "AFCI自检失败",
    "sig_midpoint_grounding": "系统接地异常",
    "sig_insulation_failure": "绝缘阻抗低",
    "sig_overcurrent": "温度过高", # 假设温度过高对应overcurrent告警
    "sig_equipment_abnormal": "设备异常", # 添加设备异常
    "sig_func_collect_fault": "功率采集器故障", # 添加功率采集器故障
    "sig_high_output_exceed": "高网输出过载" # 添加高网输出过载
}

# 处理建议映射
SUGGESTION_MAPPING = {
    "直流电弧故障": "检查对应组串路是否存在拉弧或接触不良的情况，故障排除后，需手动清除告警，再进行开机操作",
    "AFCI自检失败": "断开交流输出侧开关、直流输入侧开关，5 分钟后闭合交流输出侧开关、直流输入侧开关。如故障依然存在，请联系安装商",
    "系统接地异常": "1.请确认逆变器的保护地线是否连接正常。\n2.如果逆变器连接到TN电网，请检查N线对地电压是否正常。\n3.输入输出是否接隔离变压器，如果接入隔离变压器，设置\"接地异常关机\"为\"禁能\"",
    "绝缘阻抗低": "1.检查光伏阵列输出对保护地阻抗，如果出现短路或绝缘不足请整改故障点。\n2.检查逆变器的保护地线是否正确连接。\n3.如果强光在阴雨天环境下该阻抗确实低于默认值，请对\"绝缘阻抗保护点\"重新进行设置。",
    "温度过高": "1.检查逆变器安装位置的通风是否良好，环境温度是否超出最高允许的环境温度范围。\n2.如果不通风或环境温度过高，请改善其通风散热状况。",
    "组件碎裂": "更换组件",
    "掉串": "检查各组件间连接是否完好，检查逆变器端接入是否完好",
    "遮挡": "1.草木遮挡建议除草\n2.前后排遮挡建议调整光伏板角度",
    "积灰": "清洗组件",
    "二极管故障": "严重时更换组件",
    "热斑": "严重时更换组件"
}

# 所需工具映射
TOOLS_MAPPING = {
    "直流电弧故障": "绝缘手套、安全帽",
    "AFCI自检失败": "绝缘手套、安全帽",
    "系统接地异常": "绝缘手套、安全帽",
    "绝缘阻抗低": "电阻摇表、万用表、绝缘手套、安全帽、逆变器维修工具包",
    "温度过高": "绝缘手套、安全帽、换热风扇、逆变器维修工具包",
    "组件碎裂": "绝缘手套、安全帽、备用组件、扳手、套筒、组件更换包",
    "掉串": "绝缘手套、安全帽、MC4插头、插头更换工具",
    "遮挡": "绝缘手套、安全帽、除草设备、支架调整工具",
    "积灰": "绝缘手套、安全帽、清水、清洗组件器具",
    "二极管故障": "绝缘手套、安全帽、红外热成像仪、备用组件（如需）、扳手、套筒、组件更换包",
    "热斑": "绝缘手套、安全帽、红外热成像仪、备用组件（如需）、扳手、套筒、组件更换包"
}

def get_inverter_alarms(station_name, process_date, repo_abs_path, database_manager, station_model):
    """获取逆变器告警数据（ORM方式）"""
    try:
        # 解析日期
        date_obj = datetime.strptime(process_date, '%Y-%m-%d')
        end_time = date_obj.replace(hour=23, minute=59, second=59)
        start_time = date_obj.replace(hour=0, minute=0, second=0)
        start_timestamp = int(time.mktime(start_time.timetuple()))
        end_timestamp = int(time.mktime(end_time.timetuple()))

        # ORM模型获取
        _, InverterInfo, _ = station_model if station_model else (None, None, None)
        if InverterInfo is None:
            return []
        alarm_fields = [
            "sig_arc_failure", "sig_arc_self_protection", "sig_midpoint_grounding", "sig_insulation_failure", "sig_overcurrent",
            "sig_equipment_abnormal", "sig_func_collect_fault", "sig_high_output_exceed"
        ]
        alarm_results = []
        alarm_count = 0
        with database_manager.get_session(station_name) as session:
            query = session.query(InverterInfo).filter(
                InverterInfo.timestamp >= start_timestamp,
                InverterInfo.timestamp <= end_timestamp
            )
            results = query.all()
            for row in results:
                for field in alarm_fields:
                    if hasattr(row, field) and getattr(row, field) == 1:
                        if field in INVERTER_ALARM_MAPPING:
                            alarm_count += 1
                            alarm_name = INVERTER_ALARM_MAPPING[field]
                            alarm_results.append({
                                "order": alarm_count,
                                "alarmType": "逆变器告警",
                                "deviceCode": f"{getattr(row, 'box_id', '')}号箱变-{getattr(row, 'inverter_id', '')}号逆变器",
                                "alarmName": alarm_name,
                                "suggestion": SUGGESTION_MAPPING.get(alarm_name, ""),
                                "tools": TOOLS_MAPPING.get(alarm_name, ""),
                                "peopleCount": 2
                            })
        return alarm_results
    except Exception as e:
        return []

def get_string_alarms(station_name, process_date, repo_abs_path, processed_device_info=None):
    """获取组串告警数据"""
    try:
        # 定义告警类型与告警名称的映射
        anomaly_type_mapping = {
            "掉串": "掉串",
            "遮挡": "遮挡",
            "积灰": "积灰",
            "二极管故障": "二极管故障",
            "热斑": "热斑"
        }
        
        # 如果processed_device_info为None，初始化为空字典
        if processed_device_info is None:
            processed_device_info = {}
            
        # 用于记录已处理过的设备ID和对应的告警类型
        if station_name not in processed_device_info:
            processed_device_info[station_name] = {}
        
        alarm_results = []
        alarm_count = 0
        
        # 尝试打开JSON文件
        json_file_path = os.path.join(repo_abs_path, 'data', station_name, 'results', f'{process_date}.json')
        if not os.path.exists(json_file_path):
            return []
        
        with open(json_file_path, 'r', encoding='utf-8') as file:
            json_data = json.load(file)
            
        if 'results' not in json_data:
            return []
        
        # 处理组串告警结果
        for device_id, device_data in json_data['results'].items():
            # 检查是否有diagnosis_results字段
            if 'diagnosis_results' in device_data and len(device_data['diagnosis_results']) > 0:
                # 使用diagnosis_results中的第一个result作为告警类型
                diagnosis_result = device_data['diagnosis_results'][0]
                if 'result' in diagnosis_result and diagnosis_result['result'] in anomaly_type_mapping:
                    anomaly_type = diagnosis_result['result']
                    
                    # 拆分device_id (格式: box_id-inverter_id-string_id)
                    parts = device_id.split('-')
                    if len(parts) == 3:
                        box_id, inverter_id, string_id = parts
                        
                        alarm_count += 1
                        alarm_name = anomaly_type_mapping[anomaly_type]
                        
                        # 记录设备ID和告警类型
                        processed_device_info[station_name][device_id] = {
                            "alarm_type": alarm_name,
                            "degradation_rate": device_data.get("degradation_score", 0)
                        }
                        
                        alarm_results.append({
                            "order": alarm_count,
                            "alarmType": "组串告警",
                            "deviceCode": f"{box_id}号箱变-{inverter_id}号逆变器-{string_id}号组串",
                            "alarmName": alarm_name,
                            "suggestion": SUGGESTION_MAPPING.get(alarm_name, ""),
                            "tools": TOOLS_MAPPING.get(alarm_name, ""),
                            "peopleCount": 2 if alarm_name != "二极管故障" and alarm_name != "热斑" else 1
                        })
        
        return alarm_results
        
    except Exception as e:
        return []

def get_plan_data(station_name, process_date, repo_abs_path, database_manager, station_model):
    """获取计划数据（包括运维层面和运行层面）"""
    try:
        # 共享变量以确保两个层面处理相同的数据
        processed_device_info = {}
        
        # 获取逆变器告警数据 - 运维层面
        inverter_alarms = get_inverter_alarms(station_name, process_date, repo_abs_path, database_manager, station_model)
        silent_log(f"获取到 {len(inverter_alarms)} 个逆变器告警")
        
        # 获取组串告警数据 - 运维层面 (将处理过的设备信息存入shared_device_info)
        string_alarms = get_string_alarms(station_name, process_date, repo_abs_path, processed_device_info)
        silent_log(f"获取到 {len(string_alarms)} 个组串告警")
        
        # 获取所有场站的组串告警数据 - 运行层面
        string_alarms_all_stations = get_center_string_alarms(station_name, process_date, repo_abs_path)
        silent_log(f"获取到所有场站共 {len(string_alarms_all_stations)} 个组串告警")
        
        # 合并运维层面数据
        maintain_results = inverter_alarms + string_alarms
        
        # 按照order排序
        maintain_results.sort(key=lambda x: x["order"])
        
        # 确保order字段是连续的
        for i, item in enumerate(maintain_results):
            item["order"] = i + 1
            
        
        return {
            "maintain_results": maintain_results,
            "runtime_results": string_alarms_all_stations
        }
        
    except Exception as e:
        return {
            "maintain_results": [],
            "runtime_results": []
        }

def get_center_string_alarms(station_name, process_date, repo_abs_path):
    """获取指定场站所属运维中心下所有场站的组串告警数据
    
    参数:
        station_name: 场站名称，用于确定所属运维中心
        process_date: 处理日期，格式为YYYY-MM-DD
        repo_abs_path: 仓库绝对路径
    
    返回:
        该运维中心下所有场站的组串告警数据列表
    """
    try:
        # 转换日期格式
        date_obj = datetime.strptime(process_date, '%Y-%m-%d')
        date_str = date_obj.strftime('%Y-%m-%d')
        
        # 根据传入的场站名称找到对应的运维中心
        target_center = None
        for center, stations in CENTER_TO_STATION.items():
            if station_name in stations:
                target_center = center
                break
        
        if target_center is None:
            silent_log(f"场站 {station_name} 未找到对应的运维中心")
            return []
        
        # 获取该运维中心下的所有场站
        center_stations = CENTER_TO_STATION[target_center]
        silent_log(f"运维中心 {target_center} 包含场站: {center_stations}")
        
        # 保存结果的列表
        all_string_alarms = []
        alarm_count = 0
        
        # 遍历该运维中心下的所有场站
        for center_station_name in center_stations:
            # 组串告警JSON文件路径
            json_file_path = os.path.join(repo_abs_path, 'data', center_station_name, 'results', f'{date_str}.json')
            
            # 检查文件是否存在
            if not os.path.exists(json_file_path):
                silent_log(f"场站 {center_station_name} 的 {date_str}.json 文件不存在")
                continue
                
            try:
                # 读取JSON数据
                with open(json_file_path, 'r', encoding='utf-8') as file:
                    json_data = json.load(file)
                
                # 确保有results字段
                if 'results' not in json_data:
                    silent_log(f"场站 {center_station_name} 的 JSON 数据中没有 results 字段")
                    continue
                
                string_count = 0
                    
                # 处理组串告警数据
                for device_id, device_data in json_data['results'].items():
                    # 检查是否有diagnosis_results字段
                    if 'diagnosis_results' in device_data and len(device_data['diagnosis_results']) > 0:
                        # 获取诊断结果
                        diagnosis_result = device_data['diagnosis_results'][0]
                        
                        # 确保有result字段
                        if 'result' in diagnosis_result:
                            result_type = diagnosis_result['result']
                            
                            # 拆分device_id (格式: box_id-inverter_id-string_id)
                            parts = device_id.split('-')
                            if len(parts) == 3:
                                box_id, inverter_id, string_id = parts
                                
                                # 检查劣化率字段（兼容两种可能的字段名）
                                degradation_rate = 0
                                raw_rate = 0
                                if "degradation_score" in device_data:
                                    raw_rate = device_data.get("degradation_score", 0)
                                # 尝试将读取到的值转换为浮点数，并处理百分比字符串
                                try:
                                    if isinstance(raw_rate, str) and '%' in raw_rate:
                                        degradation_rate = float(raw_rate.strip().replace('%', '')) / 100.0
                                    else:
                                        degradation_rate = float(raw_rate)
                                except (ValueError, TypeError):
                                    degradation_rate = 0
                                
                                # 计算累计损失电量
                                loss_amount = 0
                                if "accumulated_loss" in device_data:
                                    loss_amount = round(device_data["accumulated_loss"])
                                
                                # 从JSON中读取未来一周预计损失电量
                                estimated_loss = 0
                                if "future_loss" in device_data and isinstance(device_data["future_loss"], list):
                                    # 将future_loss数组中的所有值加起来
                                    estimated_loss = round(sum(device_data["future_loss"]) / 1000)
                                
                                # 增加告警计数
                                alarm_count += 1
                                string_count += 1
                                
                                # 添加到结果列表
                                all_string_alarms.append({
                                    "order": alarm_count,
                                    "stationName": STATION_NAME_MAPPING.get(center_station_name, center_station_name),
                                    "deviceCode": f"{box_id}号箱变-{inverter_id}号逆变器-{string_id}号组串",
                                    "alarmType": result_type,
                                    "degradationRate": f"{int(degradation_rate * 100)}%" if degradation_rate > 0 else "0%",
                                    "lossAmount": f"{loss_amount}kWh",
                                    "estimatedLoss": f"{estimated_loss}kWh",
                                    "detectionTime": date_str
                                })
                
                
            except Exception as e:
                # 处理单个场站的异常，继续处理其他场站
                silent_log(f"处理场站 {center_station_name} 时出错: {str(e)}")
                continue
                
        # 按照劣化率降序排序
        all_string_alarms.sort(key=lambda x: int(x["degradationRate"].replace("%", "") or "0"), reverse=True)
        
        # 重新设置order字段
        for i, alarm in enumerate(all_string_alarms):
            alarm["order"] = i + 1
            
        return all_string_alarms
        
    except Exception as e:
        # 处理整体异常
        return []

def export_maintain_report(station_name, process_date, repo_abs_path, maintain_results):
    columns = {
        "order": "排序",
        "alarmType": "告警类型",
        "deviceCode": "生产设备编号",
        "alarmName": "告警名称",
        "suggestion": "处理建议",
        "tools": "所需工具",
        "peopleCount": "人员数量"
    }
    df = pd.DataFrame(maintain_results)
    df = df[list(columns.keys())]
    df.rename(columns=columns, inplace=True)
    # 保存到 data/station_name/reports
    station_cn = STATION_NAME_MAPPING.get(station_name, station_name)
    save_dir = os.path.join(repo_abs_path, "data", station_name, "reports")
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, f"{station_cn}_{process_date}_运维层面.xlsx")
    df.to_excel(file_path, index=False)

    # 用 openpyxl 插入图片
    wb = load_workbook(file_path)
    ws = wb.active

    # 设置指定列宽
    col_widths = {
        "生产设备编号": 35,
        "告警类型": 10,
        "告警名称": 16,
        "处理建议": 32,
        "所需工具": 18,
        "人员数量": 10,
    }
    for col in ws.iter_cols(min_row=1, max_row=1):
        header = col[0].value
        if header in col_widths:
            ws.column_dimensions[col[0].column_letter].width = col_widths[header]

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)
    return file_path

def export_report(station_name, process_date, repo_abs_path, database_manager, station_model):
    plan_data = get_plan_data(station_name, process_date, repo_abs_path, database_manager, station_model)
    maintain_results = plan_data.get("maintain_results", [])
    runtime_results = plan_data.get("runtime_results", [])

    # 导出运维层面
    export_maintain_report(station_name, process_date, repo_abs_path, maintain_results)
    # 导出运行层面
    export_runtime_report_with_images(station_name, process_date, repo_abs_path, runtime_results)

    station_cn = STATION_NAME_MAPPING.get(station_name, station_name)
    maintain_path = os.path.join(repo_abs_path, "data", station_name, "reports", f"{station_cn}_{process_date}_运维层面.xlsx")
    runtime_path = maintain_path.replace("运维层面.xlsx", "运行层面.xlsx")
    return maintain_path, runtime_path

def export_runtime_report_with_images(station_name, process_date, repo_abs_path, runtime_results):
    columns = [
        ("order", "排序"),
        ("stationName", "电站名称"),
        ("deviceCode", "生产设备编号"),
        ("alarmType", "故障类型"),
        ("lossAmount", "累计损失电量（千瓦时）"),
        ("estimatedLoss", "未来一周预计损失电量（千瓦时）"),
        ("degradationRate", "劣化率"),
        ("rgb_image", "无人机巡检可见光图像"),
        ("ir_image", "无人机巡检红外图像"),
    ]

    rows = []
    img_info = []  # 记录图片路径和行号
    img_count = 0  # 统计插入图片数量
    for idx, item in enumerate(runtime_results):
        match = re.match(r"(\d+)号箱变-(\d+)号逆变器-(\d+)号组串", item["deviceCode"])
        if match:
            id_str = f"{int(match.group(1))},{int(match.group(2))},{int(match.group(3))}"
        else:
            id_str = ""

        img_dir = os.path.join(repo_abs_path, "data", station_name, "images")
        rgb_path = os.path.join(img_dir, f"{id_str}_rgb.jpg")
        ir_path = os.path.join(img_dir, f"{id_str}_ir.jpg")

        # 记录图片路径和行号
        img_info.append((idx+2, rgb_path if os.path.exists(rgb_path) else None, ir_path if os.path.exists(ir_path) else None))

        row = [
            item.get("order", ""),
            item.get("stationName", ""),
            item.get("deviceCode", ""),
            item.get("alarmType", ""),
            item.get("lossAmount", "").replace("kWh", ""),  # 去掉kWh后缀
            item.get("estimatedLoss", "").replace("kWh", ""),  # 去掉kWh后缀
            item.get("degradationRate", ""),
            "",  # 图片列先留空
            "",
        ]
        rows.append(row)

    df = pd.DataFrame(rows, columns=[col[1] for col in columns])
    # 保存到 data/station_name/reports
    # 获取center英文名
    target_center = None
    for center, stations in CENTER_TO_STATION.items():
        if station_name in stations:
            target_center = center
            break
    # 获取center中文名
    center_cn = CENTER_NAME_MAPPING.get(target_center, target_center or station_name)
    save_dir = os.path.join(repo_abs_path, "data", station_name, "reports")
    os.makedirs(save_dir, exist_ok=True)
    file_path = os.path.join(save_dir, f"{center_cn}_{process_date}_运行层面.xlsx")
    df.to_excel(file_path, index=False)

    # 用 openpyxl 插入图片
    wb = load_workbook(file_path)
    ws = wb.active

    # 设置指定列宽
    col_widths = {
        "生产设备编号": 35,
        "故障类型": 12,
        "累计损失电量（千瓦时）": 22,
        "未来一周预计损失电量（千瓦时）": 30,
        "无人机巡检可见光图像": 20,
        "无人机巡检红外图像": 20,
    }
    for col in ws.iter_cols(min_row=1, max_row=1):
        header = col[0].value
        if header in col_widths:
            ws.column_dimensions[col[0].column_letter].width = col_widths[header]

    for row_idx, rgb_path, ir_path in img_info:
        if rgb_path:
            img = XLImage(rgb_path)
            img.width = 192  # 可根据需要调整
            img.height = 108
            ws.add_image(img, f"H{row_idx}")
            img_count += 1
        if ir_path:
            img = XLImage(ir_path)
            img.width = 135
            img.height = 108
            ws.add_image(img, f"I{row_idx}")
            img_count += 1

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)
    return file_path

def export_maintain_report_only(station_name, process_date, repo_abs_path, database_manager, station_model):
    """仅导出运维层面报告"""
    plan_data = get_plan_data(station_name, process_date, repo_abs_path, database_manager, station_model)
    maintain_results = plan_data.get("maintain_results", [])
    
    # 导出运维层面
    maintain_path = export_maintain_report(station_name, process_date, repo_abs_path, maintain_results)
    return maintain_path

def export_runtime_report_only(station_name, process_date, repo_abs_path, database_manager, station_model):
    """仅导出运行层面报告"""
    plan_data = get_plan_data(station_name, process_date, repo_abs_path, database_manager, station_model)
    runtime_results = plan_data.get("runtime_results", [])
    
    # 导出运行层面
    runtime_path = export_runtime_report_with_images(station_name, process_date, repo_abs_path, runtime_results)
    return runtime_path